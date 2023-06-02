# Source Info:
#  {
#    "Provider": "AL_ADECA",
#    "State": "AL",
#    "Name": "Alabama Department of Economic and Community Affairs - ADECA",
#    "Title": "Office of Minority Business Enterprise – ADECA",
#    "ImportedFileName": "OMBE-certified-businesses.xlsx"
#  }

# Summary of process:
#   1. If text is in bold, it is Industry
#   2. Filter out blanks in Address B:K

# Field Mapping:
#
# ╔════════════════════════════╦═══════════════════════╦════════════╗
# ║ DEST_COL_NAME              ║ SRC_COL_NAME          ║ SRC_COL_NO ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Sequence                   ║ Program Generated     ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Company Name               ║ Business Name         ║ 1          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ DBA                        ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ DUNS Number                ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Addr Line 1                ║ Address B:K           ║ 2          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Addr Line 2                ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ City                       ║ City, State, Zip      ║ 3          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ State                      ║ City, State, Zip      ║ 3          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Zip Code                   ║ City, State, Zip      ║ 3          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Website                    ║ Website URL           ║ 7          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Contact Person             ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Phone                      ║ Phone                 ║ 5          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Cell Phone                 ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Fax                        ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Email Address              ║ Email Address         ║ 6          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Gender                     ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Ethnicity                  ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Given Diversity            ║ "MWBE"                ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Primary Diversity          ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Other Diversities          ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Certifying Agency          ║ "ADECA"               ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Certification Number       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Certification Start Date   ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Certification Expiry Date ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Primary NAICS Code         ║ NAICS                 ║ 8          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Other NAICS Codes          ║ NAICS                 ║ 8          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Business Description       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 1       ║ Industry              ║ Header     ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 2       ║ County                ║ 4          ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 3       ║ Product_Service_Codes ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 4       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 5       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 6       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 7       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 8       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 9       ║                       ║            ║
# ╠════════════════════════════╬═══════════════════════╬════════════╣
# ║ Additional Details 10      ║                       ║            ║
# ╚════════════════════════════╩═══════════════════════╩════════════╝
#

import logging

import openpyxl

from src import custom_logger, empty_dict, prepare_csv_data

log = custom_logger.LogGen.logGen()
log.setLevel(level=logging.DEBUG)

provider = "AL_ADECA"
records = []


def processFile(filename):
    log.info(f'Loaded module {provider}...')
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    row_num = 0
    csv_row_num = 0
    blank_lines = 0
    industry = ''

    while True:
        row_num += 1
        if not ws.cell(row_num, 1).value:
            blank_lines += 1
            if blank_lines > 2:
                log.info(f'three blank lines after row {row_num}. End of file reached')
                break
            continue
        blank_lines = 0
        if ws.cell(row_num, 1).font.bold:
            industry = prepare_csv_data(ws.cell(row_num, 1).value)
            continue

        dict1 = empty_dict()
        csv_row_num += 1
        dict1['source'] = provider
        dict1['sequence'] = prepare_csv_data(csv_row_num)
        dict1['company'] = prepare_csv_data(ws.cell(row_num, 1).value)
        dict1['addr1'] = prepare_csv_data(ws.cell(row_num, 2).value)
        dict1['city'] = prepare_csv_data(ws.cell(row_num, 3).value, pattern=r'^[^,]*')
        dict1['state'] = prepare_csv_data(ws.cell(row_num, 3).value, pattern=r'(?<=, )[A-Z]{2}')
        dict1['zipcode'] = prepare_csv_data(ws.cell(row_num, 3).value, pattern=r'[-0-9]+$')
        dict1['website'] = prepare_csv_data(ws.cell(row_num, 7).value)
        dict1['phone'] = prepare_csv_data(ws.cell(row_num, 5).value)
        dict1['email'] = prepare_csv_data(ws.cell(row_num, 6).value)
        dict1['gdiverse'] = 'MWBE'
        dict1['cert_agency'] = 'ADECA - Alabama Dept. of Economic and Community Affairs'
        dict1['pnaics'] = prepare_csv_data(ws.cell(row_num, 8).value, pattern=r'[0-9,;]+(?=[,;])')
        dict1['onaics'] = prepare_csv_data(ws.cell(row_num, 8).value, pattern=r'(?<=[,|;])[\s\S]*$')
        dict1['addl1'] = prepare_csv_data(industry, prefix='Industry')
        dict1['addl2'] = prepare_csv_data(ws.cell(row_num, 4).value, prefix='County')
        dict1['addl3'] = prepare_csv_data(ws.cell(row_num, 9).value, prefix='Prod_Svc_Codes')

        records.append(dict1)
        # log.info(f'Record #{csv_row_num}: {dict1}')

    return records
