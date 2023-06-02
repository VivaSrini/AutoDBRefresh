# Source Info:
#
# {
#   "Provider": "CA_LACITY_DBE",
#   "State": "CA",
#   "Name": "City of Los angels, CA",
#   "Website": "https://bca.lacity.org/CertificationListings/DBETable.php",
#   "Title": "DBE/MBE/WBE Directory",
#   "ImportedFileName": "DBEMBEWBE Directory.xlsx"
# }

# Summary of process:
#   - use Pandas read_html to import .xls file (file content is a html table!)
#     into a dataframe
#   - export dataframe to csv

# Field Mapping:
#
# ╔════════════════════════════╦═══════════════════╦════════════╗
# ║ DEST_COL_NAME              ║ SRC_COL_NAME      ║ SRC_COL_NO ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Sequence                   ║ Program Generated ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Company Name               ║ Company Name      ║ 2          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ DBA                        ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ DUNS Number                ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Addr Line 1                ║ Address           ║ 4          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Addr Line 2                ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ City                       ║ Address           ║ 4          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ State                      ║ Address           ║ 4          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Zip Code                   ║ Address           ║ 4          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Website                    ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Contact Person             ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Phone                      ║ Phone             ║ 3          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Cell Phone                 ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Fax                        ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Email Address              ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Gender                     ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Ethnicity                  ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Given Diversity            ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Primary Diversity          ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Other Diversities          ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certifying Agency          ║ "BCA Los Angeles" ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certification Number       ║ CCA#              ║ 1          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certification Start Date   ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certification Expiry Date ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Primary NAICS Code         ║ NAICS             ║ 5          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Other NAICS Codes          ║ NAICS             ║ 5          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Business Description       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 1       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 2       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 3       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 4       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 5       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 6       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 7       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 8       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 9       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 10      ║                   ║            ║
# ╚════════════════════════════╩═══════════════════╩════════════╝

import logging

import openpyxl

from src import custom_logger, empty_dict, prepare_csv_data, get_primary, get_other

log = custom_logger.LogGen.logGen()
log.setLevel(level=logging.DEBUG)

provider = "CA_LACITY_DBE"
records = []


def processFile(filename):
    log.info(f'Loaded module {provider}...')
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row_num = 3

    while ws.cell(row_num, 1).value:
        dict1 = empty_dict()
        dict1['source'] = provider
        dict1['sequence'] = prepare_csv_data(row_num - 2)
        dict1['company'] = prepare_csv_data(ws.cell(row_num, 2).value)
        dict1['addr1'] = prepare_csv_data(ws.cell(row_num, 4).value)
        dict1['phone'] = prepare_csv_data(ws.cell(row_num, 3).value)
        dict1['cert_num'] = prepare_csv_data(ws.cell(row_num, 1).value)
        dict1['email'] = prepare_csv_data(ws.cell(row_num, 6).value)
        dict1['gdiverse'] = 'MWBE'
        dict1['cert_agency'] = 'City of Los Angeles BCA'
        dict1['pnaics'] = prepare_csv_data(get_primary(ws.cell(row_num, 8).value, pattern=r'[0-9,;]+(?=[,;])'))
        dict1['onaics'] = prepare_csv_data(get_other(ws.cell(row_num, 8).value, pattern=r'(?<=[,|;])[\s\S]*$'))
        records.append(dict1)
        row_num += 1

    return records
