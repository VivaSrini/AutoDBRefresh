# Source Info:
#
# {
#   "Provider": "CA_LACITY_SLB",
#   "State": "CA",
#   "Name": "City of Los angels, CA",
#   "Website": "https://bca.lacity.org/CertificationListings/SLBDirectory.php",
#   "Title": "SLB Directory",
#   "ImportedFileName": "SLB Directory.xlsx"
# }

# Summary of process:
#   - use Pandas read_html to import .xls file (file content is a html table!)
#     into a dataframe
#   - export dataframe to csv

# Field Mapping:
#
# ╔════════════════════════════╦═══════════════════╦════════════╗
# ║ DEST_COL_NAME              ║ SRC_COL_NAME      ║ SRC_COL_NO ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Sequence                   ║ Program Generated ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Company Name               ║ Company Name      ║ 2          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ DBA                        ║ DBA               ║ 3          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ DUNS Number                ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Addr Line 1                ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Addr Line 2                ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ City                       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ State                      ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Zip Code                   ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Website                    ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Contact Person             ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Phone                      ║ Phone             ║ 4          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Cell Phone                 ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Fax                        ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Email Address              ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Gender                     ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Ethnicity                  ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Given Diversity            ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Primary Diversity          ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Other Diversities          ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certifying Agency          ║ "BCA Los Angeles" ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certification Number       ║ SLB#              ║ 1          ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certification Start Date   ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Certification Expiry Date ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Primary NAICS Code         ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Other NAICS Codes          ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Business Description       ║ Decsription       ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 1       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 2       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 3       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 4       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 5       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 6       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 7       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 8       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 9       ║                   ║            ║
# ╠════════════════════════════╬═══════════════════╬════════════╣
# ║ Additional Details 10      ║                   ║            ║
# ╚════════════════════════════╩═══════════════════╩════════════╝

import logging

import openpyxl

from src import custom_logger, empty_dict, prepare_csv_data

log = custom_logger.LogGen.logGen()
log.setLevel(level=logging.DEBUG)

provider = "CA_LACITY_SLB"
records = []


def processFile(filename):
    log.info(f'Loaded module {provider}...')
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row_num = 3

    while ws.cell(row_num, 1).value:
        dict1 = empty_dict()
        dict1['source'] = provider
        dict1['sequence'] = prepare_csv_data(row_num - 2)
        dict1['company'] = prepare_csv_data(ws.cell(row_num, 2).value)
        dict1['dba'] = prepare_csv_data(ws.cell(row_num, 3).value)
        dict1['phone'] = prepare_csv_data(ws.cell(row_num, 3).value)
        dict1['email'] = prepare_csv_data(ws.cell(row_num, 6).value)
        dict1['gdiverse'] = 'SLB'
        dict1['cert_agency'] = 'City of Los Angeles BCA'
        dict1['cert_num'] = prepare_csv_data(ws.cell(row_num, 1).value)
        dict1['biz_desc'] = prepare_csv_data(ws.cell(row_num, 5).value)
        records.append(dict1)
        row_num += 1

    return records
